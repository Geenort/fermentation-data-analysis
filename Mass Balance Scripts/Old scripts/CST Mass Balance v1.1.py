'''
Potential Errors:
1) Assumption: biomass loaded effluent liquid has density of 1000 g/L
2) acetate not considered
3) Feed scale data contains seemingly random abnormalities. It's a good idea to open the raw data and manually check scale readings 
    @T_initial, T_initial+1, T_final, and T_final-1 (these points are used to calculate dilution rate).
    Select new T_initial and new T_final where scale data at these four timepoints makes sense.
'''

print("-----------------------------------------------------")

while True:
    import pandas as pd
    from openpyxl import load_workbook

    # Primary inputs from user
    run_id = 'CST' + str(input("Enter the digits from the run ID (CST'XXX'). Only enter the digits!: "))
    start_time = float(input("Enter the START time. If timepoint is not in raw data, the next timepoint will be used. "))
    end_time = float(input("Enter the END time. If timepoint is not in raw data, the previous timepoint will be used. "))
    assumed_working_mass_g = 425.0

    # Input and output file strings
    infile = '../CST Data Files/run_set_data_' + run_id + '.csv'
    outfile = '../Mass Balance Run List_PYTHON_(0.273DCW)_v0.2.xlsx'
    varfile = '../FDI Tags/variable_tag_list.csv'

    # Assumptions/Utilities
    conv_od_to_gpL = 0.273
    T_ref_degC = 25.0
    P_ref_barg = 0.0

    water_c_h_o_n_ratio_mass = [0.0, 0.1111, 0.8889, 0.0]
    biomass_c_h_o_n_ratio_mass = [0.5424, 0.0714, 0.1778, 0.2084]
    bdo_c_h_o_n_ratio_mass = [0.5333, 0.1111, 0.355555, 0.0]

    comps = ['argon', 'carbon dioxide', 'ethane', 'hydrogen', 'methane', 'nitrogen', 'oxygen', 'propane']
    comp_mw = [40, 44, 30, 2, 16, 28, 16, 44]
    c_dict = [0, 1, 2, 0, 1, 0, 0, 3]
    h_dict = [0, 0, 6, 2, 4, 0, 0, 8]
    o_dict = [0, 2, 0, 0, 0, 0, 2, 0]
    n_dict = [0, 0, 0, 0, 0, 2, 0, 0]

    # Load data into DataFrame, and Trim DF to only include data within the specified time window. 
    # Set index to ferm age (time), and don't drop column!
    # Remove 'NaN' values from raw data
    raw_data = pd.read_csv(infile)
    raw_data = raw_data[raw_data['fermentation_age_hours'] >= start_time]
    raw_data = raw_data[raw_data['fermentation_age_hours'] <= end_time]
    raw_data = raw_data.set_index('fermentation_age_hours', drop=False)
    raw_data.fillna(0.0, inplace=True)
    # raw_data = raw_data[raw_data['offgas2.rms_flow'] >= 50.0]

    # Pull tank number from raw data as a string, e.g. "02" from "T02"
    tank = raw_data.at[raw_data.index[0],'fermentor'][1:3]

    # Initialize dict with keys only (values = None)
    variable_dict = dict.fromkeys(['Average BDO, mg/L', 'Average OD, raw', 'Feed Scale, grams', 'F_in, sccm', 'x_argon_in_raw', 'x_carbon dioxide_in_raw', 'x_ethane_in_raw', 'x_hydrogen_in_raw', 'x_methane_in_raw', 'x_nitrogen_in_raw', 'x_oxygen_in_raw', 'x_propane_in_raw', 'x_argon_out_raw', 'x_carbon dioxide_out_raw', 'x_ethane_out_raw', 'x_hydrogen_out_raw', 'x_methane_out_raw', 'x_nitrogen_out_raw', 'x_oxygen_out_raw', 'x_propane_out_raw'])

    # Fill in dict values with header tags for input tank, from varfile csv
    var_map = pd.read_csv(varfile, index_col='Tank')
    for key in variable_dict.keys():
        variable_dict[key] = var_map.at[int(tank), key]

    # Set up results DF. If exact input start_time/end_time don't exist, next/previous timepoints will be picked instead
    results = pd.DataFrame(index=[], columns=[])
    results.at[run_id, 'Run_ID'] = run_id
    results.at[run_id, 'Tank'] = int(tank)
    results.at[run_id, 'Start Time, hours'] = raw_data['fermentation_age_hours'].min()
    results.at[run_id, 'End Time, hours'] = raw_data['fermentation_age_hours'].max()
    start_time = raw_data['fermentation_age_hours'].min()
    end_time = raw_data['fermentation_age_hours'].max()

    # Warn if steady state is not reached. Steady state being +/- 2% of BDO, OD concentrations
    bdo_start = raw_data.at[start_time, variable_dict['Average BDO, mg/L']]
    bdo_end = raw_data.at[end_time, variable_dict['Average BDO, mg/L']]
    od_start = raw_data.at[start_time, variable_dict['Average OD, raw']]
    od_end = raw_data.at[end_time, variable_dict['Average OD, raw']]
    if (abs(bdo_end - bdo_start) / bdo_end) >= 0.25 or (abs(od_end - od_start) / od_end) >= 0.02:
        print(f'WARNING: This may not be a steady state time interval! (Delta BDO >= 25% or Delta OD >= 2%)')

    print("-----------------------------------------------------")
    print('Run Number:'.ljust(30) + f'{run_id}')
    print('Tank:'.ljust(30) + f'T{tank}')
    print('Start time:'.ljust(30) + f'{start_time}')
    print('End time:'.ljust(30) + f'{end_time}')
    print('Initial BDO in mg/L:'.ljust(30) + f'{bdo_start}')
    print('Final BDO in mg/L'.ljust(30) + f'{bdo_end}')
    print('Initial OD:'.ljust(30) + f'{od_start}')
    print('Final OD:'.ljust(30) + f'{od_end}')

    results.at[run_id, 'Initial BDO, mg/L'] = bdo_start
    results.at[run_id, 'Final BDO, mg/L'] = bdo_end
    results.at[run_id, 'Initial OD, raw'] = od_start
    results.at[run_id, 'Final OD, raw'] = od_start

    # Load averaged data into results DF
    for var, tag in variable_dict.items():
        # Check units on tank scale, ensure in grams
        if var == 'Media Scale, grams':
            avg = raw_data[tag].mean()
            if avg < 1000.00:
                results.at[run_id, var] = avg * 1000.0
            else:
                results.at[run_id, var] = avg
        # For remainder of variables, just take average
        else:
            results.at[run_id, var] = raw_data[tag].mean()

    ## Inlet gas
    results.at[run_id, 'Inlet Gas Rate, mol/hr'] = results.at[run_id, 'F_in, sccm'] * (1/1000) * 60 * (P_ref_barg + 1.0) / (0.08314 * (273 + T_ref_degC))
    # Normalize inlet gas composition, calculate MW
    mass_spec_sum = 0.0
    for x in range(0, len(comps)):
        tag = 'x_' + comps[x] + '_in_raw'
        mass_spec_sum += results.at[run_id, tag]
    MW_sum = 0.0
    for x in range(0, len(comps)):
        tag_raw = 'x_' + comps[x] + '_in_raw'
        tag_norm = 'x_'+ comps[x] +'_in_normalized'
        results.at[run_id, tag_norm] = results.at[run_id, tag_raw] * 100.00 / mass_spec_sum
        MW_sum += results.at[run_id, tag_norm] * comp_mw[x] / 100.0
    results.at[run_id, 'Inlet Gas MW'] = MW_sum
    # Calculate inlet gas component rates in mol/hr, elemental rates in mol/hr, elemental rates in grams/hr
    in_gas_C_rate_mph = 0.0
    in_gas_H_rate_mph = 0.0
    in_gas_O_rate_mph = 0.0
    in_gas_N_rate_mph = 0.0
    for x in range(0, len(comps)):
        tag_norm = 'x_' + comps[x] + '_in_normalized'
        tag_rate = 'F_' + comps[x] + '_in_molph'
        results.at[run_id, tag_rate] = results.at[run_id, 'Inlet Gas Rate, mol/hr'] * results.at[run_id, tag_norm] / 100.0
        in_gas_C_rate_mph += results.at[run_id, tag_rate] * c_dict[x]
        in_gas_H_rate_mph += results.at[run_id, tag_rate] * h_dict[x]
        in_gas_O_rate_mph += results.at[run_id, tag_rate] * o_dict[x]
        in_gas_N_rate_mph += results.at[run_id, tag_rate] * n_dict[x]
    results.at[run_id, 'Inlet C Rate, mol/hr'] = in_gas_C_rate_mph
    results.at[run_id, 'Inlet H Rate, mol/hr'] = in_gas_H_rate_mph
    results.at[run_id, 'Inlet O Rate, mol/hr'] = in_gas_O_rate_mph
    results.at[run_id, 'Inlet N Rate, mol/hr'] = in_gas_N_rate_mph

    results.at[run_id, 'Inlet C Rate, g/hr'] = in_gas_C_rate_mph * 12.0
    results.at[run_id, 'Inlet H Rate, g/hr'] = in_gas_H_rate_mph * 1.0
    results.at[run_id, 'Inlet O Rate, g/hr'] = in_gas_O_rate_mph * 16.0
    results.at[run_id, 'Inlet N Rate, g/hr'] = in_gas_N_rate_mph * 14.0

    ## Effluent Gas
    # Normalize inlet gas composition, calculate MW
    mass_spec_sum = 0.0
    for x in range(0, len(comps)):
        tag = 'x_' + comps[x] + '_out_raw'
        mass_spec_sum += results.at[run_id, tag]
    MW_sum = 0.0
    for x in range(0, len(comps)):
        tag_raw = 'x_' + comps[x] + '_out_raw'
        tag_norm = 'x_' + comps[x] + '_out_normalized'
        results.at[run_id, tag_norm] = results.at[run_id, tag_raw] * 100.00 / mass_spec_sum
        MW_sum += results.at[run_id, tag_norm] * comp_mw[x] / 100.0
    results.at[run_id, 'Effluent Gas MW'] = MW_sum
    # Calculate effluent flow rate in mols/hr
    results.at[run_id, 'Effluent Gas Rate, mol/hr'] = results.at[run_id, 'Inlet Gas Rate, mol/hr'] * results.at[run_id, 'x_argon_in_normalized'] / results.at[run_id, 'x_argon_out_normalized']
    # Calculate effluent gas component rates in mol/hr, elemental rates in mol/hr, elemental rates in grams/hr
    out_gas_C_rate_mph = 0.0
    out_gas_H_rate_mph = 0.0
    out_gas_O_rate_mph = 0.0
    out_gas_N_rate_mph = 0.0
    for x in range(0, len(comps)):
        tag_norm = 'x_' + comps[x] + '_out_normalized'
        tag_rate = 'F_' + comps[x] + '_out_molph'
        results.at[run_id, tag_rate] = results.at[run_id, 'Effluent Gas Rate, mol/hr'] * results.at[run_id, tag_norm] / 100.0
        out_gas_C_rate_mph += results.at[run_id, tag_rate] * c_dict[x]
        out_gas_H_rate_mph += results.at[run_id, tag_rate] * h_dict[x]
        out_gas_O_rate_mph += results.at[run_id, tag_rate] * o_dict[x]
        out_gas_N_rate_mph += results.at[run_id, tag_rate] * n_dict[x]
    results.at[run_id, 'Effluent C Rate, mol/hr'] = out_gas_C_rate_mph
    results.at[run_id, 'Effluent H Rate, mol/hr'] = out_gas_H_rate_mph
    results.at[run_id, 'Effluent O Rate, mol/hr'] = out_gas_O_rate_mph
    results.at[run_id, 'Effluent N Rate, mol/hr'] = out_gas_N_rate_mph

    results.at[run_id, 'Effluent C Rate, g/hr'] = out_gas_C_rate_mph * 12.0
    results.at[run_id, 'Effluent H Rate, g/hr'] = out_gas_H_rate_mph * 1.0
    results.at[run_id, 'Effluent O Rate, g/hr'] = out_gas_O_rate_mph * 16.0
    results.at[run_id, 'Effluent N Rate, g/hr'] = out_gas_N_rate_mph * 14.0


    ## Liquid Feed
    # Liquid feed rate is d(feed scale)/dt. This mass balance is designed for a single dilution rate (feed rate / V)
    # Step 1: check to see if rate at beginning and end of time period are the same
    t_start = start_time
    t_start_plus1 = start_time + 1.0
    t_end = end_time
    t_end_minus1 = end_time - 1.0

    # Feed rate = (feed scale reading @ t - feed scale reading @ t+1) / 1 hour
    feed_rate_start = raw_data.at[t_start, variable_dict['Feed Scale, grams']] - raw_data.at[t_start_plus1, variable_dict['Feed Scale, grams']]
    feed_rate_end = raw_data.at[t_end_minus1, variable_dict['Feed Scale, grams']] - raw_data.at[t_end, variable_dict['Feed Scale, grams']]

    delta_feed_rate = abs(((feed_rate_end - feed_rate_start) / feed_rate_end))
    if delta_feed_rate >= 0.05:
        print(f"WARNING: Dilution rate is not consistent over this time interval! Delta Feed Rate = {round(delta_feed_rate*100.00,2)}%; >= 5%)")
    # Check units. should be either grams or kg. change to grams if given in kg
    if feed_rate_end < 1.0:
        results.at[run_id, 'Liquid Feed Rate, g/hr'] = feed_rate_end * 1000.00
        results.at[run_id, 'Dilution Rate, hrs^-1'] = results.at[run_id, 'Liquid Feed Rate, g/hr'] / assumed_working_mass_g
    else:
        results.at[run_id, 'Liquid Feed Rate, g/hr'] = 0.5 * (feed_rate_start + feed_rate_end)
        results.at[run_id, 'Dilution Rate, hrs^-1'] = results.at[run_id, 'Liquid Feed Rate, g/hr'] / assumed_working_mass_g

    results.at[run_id, 'Liquid C Rate, g/hr'] = results.at[run_id, 'Liquid Feed Rate, g/hr'] * water_c_h_o_n_ratio_mass[0]
    results.at[run_id, 'Liquid H Rate, g/hr'] = results.at[run_id, 'Liquid Feed Rate, g/hr'] * water_c_h_o_n_ratio_mass[1]
    results.at[run_id, 'Liquid O Rate, g/hr'] = results.at[run_id, 'Liquid Feed Rate, g/hr'] * water_c_h_o_n_ratio_mass[2]
    results.at[run_id, 'Liquid N Rate, g/hr'] = results.at[run_id, 'Liquid Feed Rate, g/hr'] * water_c_h_o_n_ratio_mass[3]

    ## Product Generation
    # Start with accumulation. Accumulation of C = Mass * (delta_w_bdo * sigma_O_bdo + delta_w_bm * sigma_O_bm + delta_w_h2o * sigma_O_h2o)
    delta_w_bdo_gpL = (bdo_end - bdo_start) / 1000.0
    delta_w_bm_gpL = (od_end - od_end) * conv_od_to_gpL
    delta_w_h2o_gpL = -(delta_w_bdo_gpL + delta_w_bm_gpL)
    results.at[run_id, 'Accumulation C, g/hr'] = (assumed_working_mass_g / 1000.0 ) * ( delta_w_bdo_gpL * bdo_c_h_o_n_ratio_mass[0] + \
                                                                                delta_w_bm_gpL * biomass_c_h_o_n_ratio_mass[0] + \
                                                                                delta_w_h2o_gpL * water_c_h_o_n_ratio_mass[0] ) \
                                        / (end_time - start_time)
    results.at[run_id, 'Accumulation H, g/hr'] = (assumed_working_mass_g / 1000.0 ) * ( delta_w_bdo_gpL * bdo_c_h_o_n_ratio_mass[1] + \
                                                                                delta_w_bm_gpL * biomass_c_h_o_n_ratio_mass[1] + \
                                                                                delta_w_h2o_gpL * water_c_h_o_n_ratio_mass[1] ) \
                                        / (end_time - start_time)
    results.at[run_id, 'Accumulation O, g/hr'] = (assumed_working_mass_g / 1000.0 ) * ( delta_w_bdo_gpL * bdo_c_h_o_n_ratio_mass[2] + \
                                                                                delta_w_bm_gpL * biomass_c_h_o_n_ratio_mass[2] + \
                                                                                delta_w_h2o_gpL * water_c_h_o_n_ratio_mass[2] ) \
                                        / (end_time - start_time)
    results.at[run_id, 'Accumulation N, g/hr'] = (assumed_working_mass_g / 1000.0 ) * ( delta_w_bdo_gpL * bdo_c_h_o_n_ratio_mass[3] + \
                                                                                delta_w_bm_gpL * biomass_c_h_o_n_ratio_mass[3] + \
                                                                                delta_w_h2o_gpL * water_c_h_o_n_ratio_mass[3] ) \
                                        / (end_time - start_time)

    # Determine liquid mass fraction water
    results.at[run_id, 'w_BDO_liq_g/L'] = results.at[run_id, 'Average BDO, mg/L'] / 1000.0
    results.at[run_id, 'w_BM_liq_g/L'] = results.at[run_id, 'Average OD, raw'] * conv_od_to_gpL
    results.at[run_id, 'w_H2O_liq_g/L'] = 1000.0 - results.at[run_id, 'w_BDO_liq_g/L'] - results.at[run_id, 'w_BM_liq_g/L']
    # Oxygen balance. In - Out = Acc, so m_in_O_gas + m_in_O_liq - m_acc_O = m_out_O_gas + m_out_O_liq
    # m_out_O_liq = v_out_liq * (w_BDO_liq * w_O_BDO + w_BM_liq * w_O_BDO + w_H2O_liq * w_O_H2O)
    # Another option is to simply set 'Draw Rate, g/hr' = 'Liquid Feed Rate, g/hr'
    results.at[run_id, 'Draw Rate, L/hr'] = (results.at[run_id, 'Liquid O Rate, g/hr'] + results.at[run_id, 'Inlet O Rate, g/hr'] - results.at[run_id, 'Effluent O Rate, g/hr'] - results.at[run_id, 'Accumulation O, g/hr']) \
        / (results.at[run_id, 'w_BDO_liq_g/L'] * bdo_c_h_o_n_ratio_mass[2] + results.at[run_id, 'w_BM_liq_g/L'] * biomass_c_h_o_n_ratio_mass[2] + results.at[run_id, 'w_H2O_liq_g/L'] * water_c_h_o_n_ratio_mass[2])
    results.at[run_id, 'Draw Rate, g/hr'] = results.at[run_id, 'Draw Rate, L/hr'] *1000.0

    # Determine liquid draw elemental rates
    results.at[run_id, 'Draw C Rate, g/hr'] = results.at[run_id, 'Draw Rate, L/hr'] * (results.at[run_id, 'w_BDO_liq_g/L'] * bdo_c_h_o_n_ratio_mass[0] + \
                                                                                        results.at[run_id, 'w_BM_liq_g/L'] * biomass_c_h_o_n_ratio_mass[0] + \
                                                                                        results.at[run_id, 'w_H2O_liq_g/L'] * water_c_h_o_n_ratio_mass[0])
    results.at[run_id, 'Draw H Rate, g/hr'] = results.at[run_id, 'Draw Rate, L/hr'] * (results.at[run_id, 'w_BDO_liq_g/L'] * bdo_c_h_o_n_ratio_mass[1] + \
                                                                                        results.at[run_id, 'w_BM_liq_g/L'] * biomass_c_h_o_n_ratio_mass[1] + \
                                                                                        results.at[run_id, 'w_H2O_liq_g/L'] * water_c_h_o_n_ratio_mass[1])
    results.at[run_id, 'Draw O Rate, g/hr'] = results.at[run_id, 'Draw Rate, L/hr'] * (results.at[run_id, 'w_BDO_liq_g/L'] * bdo_c_h_o_n_ratio_mass[2] + \
                                                                                        results.at[run_id, 'w_BM_liq_g/L'] * biomass_c_h_o_n_ratio_mass[2] + \
                                                                                        results.at[run_id, 'w_H2O_liq_g/L'] * water_c_h_o_n_ratio_mass[2])
    results.at[run_id, 'Draw N Rate, g/hr'] = results.at[run_id, 'Draw Rate, L/hr'] * (results.at[run_id, 'w_BDO_liq_g/L'] * bdo_c_h_o_n_ratio_mass[3] + \
                                                                                        results.at[run_id, 'w_BM_liq_g/L'] * biomass_c_h_o_n_ratio_mass[3] + \
                                                                                        results.at[run_id, 'w_H2O_liq_g/L'] * water_c_h_o_n_ratio_mass[3])

    ## Calculate Closures
    results.at[run_id, 'C Closure'] = ( results.at[run_id, 'Effluent C Rate, g/hr'] + results.at[run_id, 'Draw C Rate, g/hr'] + results.at[run_id, 'Accumulation C, g/hr'] ) /\
                                    ( results.at[run_id, 'Inlet C Rate, g/hr'] + results.at[run_id, 'Liquid C Rate, g/hr'] )
    results.at[run_id, 'H Closure'] = ( results.at[run_id, 'Effluent H Rate, g/hr'] + results.at[run_id, 'Draw H Rate, g/hr'] + results.at[run_id, 'Accumulation H, g/hr'] ) /\
                                    ( results.at[run_id, 'Inlet H Rate, g/hr'] + results.at[run_id, 'Liquid H Rate, g/hr'] )
    results.at[run_id, 'O Closure'] = ( results.at[run_id, 'Effluent O Rate, g/hr'] + results.at[run_id, 'Draw O Rate, g/hr'] + results.at[run_id, 'Accumulation O, g/hr'] ) /\
                                    ( results.at[run_id, 'Inlet O Rate, g/hr'] + results.at[run_id, 'Liquid O Rate, g/hr'] )
    results.at[run_id, 'N Closure'] = ( results.at[run_id, 'Effluent N Rate, g/hr'] + results.at[run_id, 'Draw N Rate, g/hr'] + results.at[run_id, 'Accumulation N, g/hr'] ) /\
                                    ( results.at[run_id, 'Inlet N Rate, g/hr'] + results.at[run_id, 'Liquid N Rate, g/hr'] )

    print(f"Carbon Closure: {results.at[run_id, 'C Closure']}")

    # Order output so that 'Closure' cols come first
    beginning_cols = ['Run_ID', 'Tank', 'Start Time, hours', 'End Time, hours', 'C Closure', 'H Closure', 'O Closure', 'N Closure']
    reorganized_df = results[[col for col in beginning_cols] + [col for col in results if col not in beginning_cols]]

    # Append to first unoccupied row of 'Python Raw' worksheet in outfile
    with pd.ExcelWriter(outfile, engine='openpyxl', mode='a') as writer:
        book = load_workbook(outfile)
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}
        reorganized_df.to_excel(writer, sheet_name='Python Raw', startrow=writer.sheets['Python Raw'].max_row, index=False, header=False)

    print("-----------------------------------------------------")

    # Code for restarting script upon user input
    while True:
        answer = input("Would you like run again? Enter 'y' to run again, or 'n' to continue: ")
        if answer.lower() in ('y', 'n'):
            break
        print("Invalid input. Please enter 'y' or 'n'.")
    if answer.lower() == 'y':
        print("-----------------------------------------------------")
        continue
    else:
        break