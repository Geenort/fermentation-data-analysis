'''
Potential Errors:
1) Assumption: biomass loaded effluent liquid has density of 1000 g/L
2) acetate not considered
3) Feed scale data contains seemingly random abnormalities. It's a good idea to open the raw data and manually check scale readings 
    @T_initial, T_initial+1, T_final, and T_final-1 (these points are used to calculate dilution rate).
    Select new T_initial and new T_final where scale data at these four timepoints makes sense.
'''

print("-----------------------------------------------------")

# While loop wrapper to enable re-running script
while True:
    # exception handling for FileNotFoundError
    try:
        import pandas as pd
        from openpyxl import load_workbook

        # Primary inputs from user
        run_id = 'CST' + str(input("Enter the digits from the run ID (CST'XXX'). Only enter the digits!: "))
        start_time = float(input("Enter the START time. If timepoint is not in raw data, the next timepoint will be used. "))
        end_time = float(input("Enter the END time. If timepoint is not in raw data, the previous timepoint will be used. "))

        # Input and output file strings
        infile = '../CST Data Files/run_set_data_' + run_id + '.csv'
        outfile = '../Non-Blend Mass Balance Run List PYTHON v0.4.xlsx'
        varfile = 'FDI Tags (do not touch)/tank_tags_map.csv'

        # Assumptions/Utilities
        OD_to_gDCW_perL = 0.300
        T_ref_degC = 25.0
        P_ref_barg = 0.0
        assumed_working_mass_g = 425.0

        water_c_h_o_n_ratio_mass = [0.0, 0.1111, 0.8889, 0.0]
        biomass_c_h_o_n_ratio_mass = [0.5424, 0.0714, 0.1778, 0.2084]
        bdo_c_h_o_n_ratio_mass = [0.5333, 0.1111, 0.355555, 0.0]

        # Dictionary to hold information for each of 8 components
        components_ORIGINAL = [{'name':'argon',          'molar_mass':40, 'C':0, 'H':0, 'O':0, 'N':0},
                               {'name':'carbon_dioxide', 'molar_mass':44, 'C':1, 'H':0, 'O':2, 'N':0},
                               {'name':'ethane',         'molar_mass':30, 'C':2, 'H':6, 'O':0, 'N':0},
                               {'name':'hydrogen',       'molar_mass':2,  'C':0, 'H':2, 'O':0, 'N':0},
                               {'name':'methane',        'molar_mass':16, 'C':1, 'H':4, 'O':0, 'N':0},
                               {'name':'nitrogen',       'molar_mass':28, 'C':0, 'H':0, 'O':0, 'N':2},
                               {'name':'oxygen',         'molar_mass':16, 'C':0, 'H':0, 'O':2, 'N':0},
                               {'name':'propane',        'molar_mass':44, 'C':3, 'H':8, 'O':0, 'N':0}
        ]

        components = [component for component in components_ORIGINAL if component['name'] != 'ethane' and component['name'] != 'propane']

        # Load data into DataFrame, and Trim DF to only include data within the specified time window. 
        # Set index to ferm age (time), and don't drop column!
        # Remove 'NaN' values from raw data
        raw_data = pd.read_csv(infile)
        raw_data = raw_data[raw_data['fermentation_age_hours'] >= start_time]
        raw_data = raw_data[raw_data['fermentation_age_hours'] <= end_time]
        raw_data = raw_data.set_index('fermentation_age_hours', drop=False)
        raw_data.fillna(0.0, inplace=True)
        # raw_data = raw_data[raw_data['offgas2.rms_flow'] >= 50.0]

        # Pull tank number from raw data as a string, e.g. "02" from "T02"
        tank = raw_data.at[raw_data.index[0],'fermentor'][1:3]

        # Initialize dict with keys only (values = None)
        variable_dict_ORIGINAL = dict.fromkeys(['BDO, mg/L', 'OD, raw', 'Feed Scale, kg', 
                                    'BLEND_F_in', 'BLEND_x_argon_in_raw', 'BLEND_x_carbon_dioxide_in_raw', 'BLEND_x_ethane_in_raw', 'BLEND_x_hydrogen_in_raw', 'BLEND_x_methane_in_raw', 'BLEND_x_nitrogen_in_raw', 'BLEND_x_oxygen_in_raw', 'BLEND_x_propane_in_raw',
                                    'AIR_F_in', 'AIR_x_argon_in_raw', 'AIR_x_carbon_dioxide_in_raw', 'AIR_x_ethane_in_raw', 'AIR_x_hydrogen_in_raw', 'AIR_x_methane_in_raw', 'AIR_x_nitrogen_in_raw', 'AIR_x_oxygen_in_raw', 'AIR_x_propane_in_raw',
                                    'CH4_F_in', 'CH4_x_argon_in_raw', 'CH4_x_carbon_dioxide_in_raw', 'CH4_x_ethane_in_raw', 'CH4_x_hydrogen_in_raw', 'CH4_x_methane_in_raw', 'CH4_x_nitrogen_in_raw', 'CH4_x_oxygen_in_raw', 'CH4_x_propane_in_raw',
                                    'CNG_F_in', 'CNG_x_argon_in_raw', 'CNG_x_carbon_dioxide_in_raw', 'CNG_x_ethane_in_raw', 'CNG_x_hydrogen_in_raw', 'CNG_x_methane_in_raw', 'CNG_x_nitrogen_in_raw', 'CNG_x_oxygen_in_raw', 'CNG_x_propane_in_raw',
                                    'O2_F_in', 'O2_x_argon_in_raw', 'O2_x_carbon_dioxide_in_raw', 'O2_x_ethane_in_raw', 'O2_x_hydrogen_in_raw', 'O2_x_methane_in_raw', 'O2_x_nitrogen_in_raw', 'O2_x_oxygen_in_raw', 'O2_x_propane_in_raw',
                                    'N2_F_in', 'N2_x_argon_in_raw', 'N2_x_carbon_dioxide_in_raw', 'N2_x_ethane_in_raw', 'N2_x_hydrogen_in_raw', 'N2_x_methane_in_raw', 'N2_x_nitrogen_in_raw', 'N2_x_oxygen_in_raw', 'N2_x_propane_in_raw',
                                    'x_argon_out_raw', 'x_carbon_dioxide_out_raw', 'x_ethane_out_raw', 'x_hydrogen_out_raw', 'x_methane_out_raw', 'x_nitrogen_out_raw', 'x_oxygen_out_raw', 'x_propane_out_raw'
        ])

        # v1.4b: _ethane_raw and _propane_raw were not scheduled on offgas1 mass spec from ~07/18/19 to 08/19/19
        # removed keys with _ethane or _propane from variable dict
        variable_dict = {k: v for k, v in variable_dict_ORIGINAL.items() if '_ethane' not in k and '_propane' not in k}

        # Fill in dict values with header tags for input tank, from varfile csv
        var_map = pd.read_csv(varfile, index_col='Tank')
        for key in variable_dict.keys():
            variable_dict[key] = var_map.at[int(tank), key]


        # Read current run list into DataFrame. Append new empty row, and set idx to integer index of this last row
        results = pd.read_excel(outfile, index_col=0)
        results = results.append(pd.Series(), ignore_index=True)
        idx = results.index[-1]

        results.at[idx, 'Run_ID'] = run_id
        results.at[idx, 'Tank'] = int(tank)
        results.at[idx, 'DCW, g/L/OD'] = OD_to_gDCW_perL
        results.at[idx, 'Assumed Working Vol, L'] = assumed_working_mass_g / 1000

        # If exact input start_time/end_time don't exist, next/previous timepoints will be picked instead
        results.at[idx, 'Start Time, hours'] = raw_data['fermentation_age_hours'].min()
        results.at[idx, 'End Time, hours'] = raw_data['fermentation_age_hours'].max()
        start_time = raw_data['fermentation_age_hours'].min()
        end_time = raw_data['fermentation_age_hours'].max()

        # Warn if steady state is not reached. Steady state being +/- 25% of BDO, 2% of OD
        bdo_start = raw_data.at[start_time, variable_dict['BDO, mg/L']]
        bdo_end = raw_data.at[end_time, variable_dict['BDO, mg/L']]
        od_start = raw_data.at[start_time, variable_dict['OD, raw']]
        od_end = raw_data.at[end_time, variable_dict['OD, raw']]
        if (abs(bdo_end - bdo_start) / bdo_end) >= 0.25 or (abs(od_end - od_start) / od_end) >= 0.02:
            print(f'WARNING: This may not be a steady state time interval! (Delta BDO >= 25% or Delta OD >= 2%)')

        print("-----------------------------------------------------")
        print('Run Number:'.ljust(30) + f'{run_id}')
        print('Tank:'.ljust(30) + f'T{tank}')
        print('Start time:'.ljust(30) + f'{start_time}')
        print('End time:'.ljust(30) + f'{end_time}')
        print('Initial BDO in mg/L:'.ljust(30) + f'{bdo_start}')
        print('Final BDO in mg/L'.ljust(30) + f'{bdo_end}')
        print('Initial OD:'.ljust(30) + f'{od_start}')
        print('Final OD:'.ljust(30) + f'{od_end}')

        results.at[idx, 'Initial BDO, mg/L'] = bdo_start
        results.at[idx, 'Final BDO, mg/L'] = bdo_end
        results.at[idx, 'Initial OD, raw'] = od_start
        results.at[idx, 'Final OD, raw'] = od_end

        # Load averaged data into results DF
        for var, tag in variable_dict.items():
            results.at[idx, var] = raw_data[tag].mean()

        ## Inlet gas

        # List of all MFCs / streams considered in calculations
        used_MFCs = ['AIR', 'CH4', 'CNG']
        print('Gas streams considered in calculations: ' + f'{used_MFCs}')
       
        # Normalize inlet gas compositions and calculate total inlet molar weight
        # 8 components * 3 MFC's - add up 24x component fractions (averaged over time period, given in % points).
        # Calculate sum of all component fractions
        raw_comp_sum = 0.0
        for component in components:
            for ref in used_MFCs:
                tag = f'{ref}_x_' + component['name'] + '_in_raw'
                
                # raw_comp_sum should be almost # of MFCs x 100 % points
                raw_comp_sum += results.at[idx, tag]

        total_molar_weight = 0.0
        for component in components:
            for ref in used_MFCs:
                tag = f'{ref}_x_' + component['name'] + '_in_raw'
                tag_norm = f'{ref}_x_'+ component['name'] +'_in_normalized'
                
                # Normalize compositions so that each of # of MFCs x 8 component fractions add up to 100 % points
                # (ignore other negligible components)
                results.at[idx, tag_norm] = results.at[idx, tag] * len(used_MFCs) * 100.0 / raw_comp_sum
                
                # Calculate and output total Inlet molar weight
                total_molar_weight += results.at[idx, tag] * component['molar_mass'] / raw_comp_sum

        results.at[idx, 'Inlet Gas MW, g/mol'] = total_molar_weight

        # Convert total volumetric flow rate through each MFC (SCCM/min = standard cubic centimeters/min; same as mL/min)
        # to total molar flow rates (mol/hr) using PV=nRT. Then sum all 3 to get total inlet molar flow rate
        total_inlet_mph = 0.0
        for ref in used_MFCs:
            results.at[idx, f'{ref} Inlet Gas Rate, mol/hr'] = results.at[idx, f'{ref}_F_in'] * (1/1000) * 60 * (P_ref_barg + 1.0) / (0.08314 * (273 + T_ref_degC))
            total_inlet_mph += results.at[idx, f'{ref} Inlet Gas Rate, mol/hr']
            
        results.at[idx, 'Total Inlet Gas Rate, mol/hr'] = total_inlet_mph

        # Calculate elemental molar flow rates
        in_gas_C_rate_mph = 0.0
        in_gas_H_rate_mph = 0.0
        in_gas_O_rate_mph = 0.0
        in_gas_N_rate_mph = 0.0

        for component in components:
            for ref in used_MFCs:
                tag_norm = f'{ref}_x_'+ component['name'] +'_in_normalized'
                tag_rate_mph = f'{ref}_F_' + component['name'] + '_in_molph'
                
                # Calculate each of 24 component molar flow rates. MFC flow rate * component fraction / 100%
                results.at[idx, tag_rate_mph] = results.at[idx, f'{ref} Inlet Gas Rate, mol/hr'] * results.at[idx, tag_norm] / 100.0
                
                # Calculate elemental molar flow rate, sum to get total
                in_gas_C_rate_mph += results.at[idx, tag_rate_mph] * component['C']
                in_gas_H_rate_mph += results.at[idx, tag_rate_mph] * component['H']
                in_gas_O_rate_mph += results.at[idx, tag_rate_mph] * component['O']
                in_gas_N_rate_mph += results.at[idx, tag_rate_mph] * component['N']
                
        results.at[idx, 'Inlet C Rate, mol/hr'] = in_gas_C_rate_mph
        results.at[idx, 'Inlet H Rate, mol/hr'] = in_gas_H_rate_mph
        results.at[idx, 'Inlet O Rate, mol/hr'] = in_gas_O_rate_mph
        results.at[idx, 'Inlet N Rate, mol/hr'] = in_gas_N_rate_mph

        # Calculate elemental mass flow rate
        results.at[idx, 'Inlet C Rate, g/hr'] = in_gas_C_rate_mph * 12.0
        results.at[idx, 'Inlet H Rate, g/hr'] = in_gas_H_rate_mph * 1.0
        results.at[idx, 'Inlet O Rate, g/hr'] = in_gas_O_rate_mph * 16.0
        results.at[idx, 'Inlet N Rate, g/hr'] = in_gas_N_rate_mph * 14.0


        ## Effluent Gas - same steps as for Inlet Gas, but only one out stream, with same 8 components

        raw_comp_sum = 0.0
        for component in components:
            tag = 'x_' + component['name'] + '_out_raw'
            raw_comp_sum += results.at[idx, tag]

        total_molar_weight = 0.0
        for component in components:
            tag = 'x_' + component['name'] + '_out_raw'
            tag_norm = 'x_' + component['name'] + '_out_normalized'
            results.at[idx, tag_norm] = results.at[idx, tag] * 100.00 / raw_comp_sum
            total_molar_weight += results.at[idx, tag] * component['molar_mass'] / raw_comp_sum
            
        results.at[idx, 'Effluent Gas MW, g/mol'] = total_molar_weight

        # Need to calculate total effluent molar flow rate using ratio of Argon in/out, (argon is non-reactive so is not consumed)
        # = (total inlet flow) * (Argon fraction in) / (Argon fraction out)
        # Inlet flow comes from 3 MFC's; add molar flow rates from each MFC to get total

        results.at[idx, 'Effluent Gas Rate, mol/hr'] = \
        (results.at[idx, 'AIR Inlet Gas Rate, mol/hr'] * results.at[idx, 'AIR_x_argon_in_normalized'] + \
        results.at[idx, 'CH4 Inlet Gas Rate, mol/hr'] * results.at[idx, 'CH4_x_argon_in_normalized'] + \
        results.at[idx, 'CNG Inlet Gas Rate, mol/hr'] * results.at[idx, 'CNG_x_argon_in_normalized']) / \
        results.at[idx,'x_argon_out_normalized']

        out_gas_C_rate_mph = 0.0
        out_gas_H_rate_mph = 0.0
        out_gas_O_rate_mph = 0.0
        out_gas_N_rate_mph = 0.0

        for component in components:
            tag_norm = 'x_' + component['name'] + '_out_normalized'
            tag_rate_mph = 'F_' + component['name'] + '_out_molph'
            
            results.at[idx, tag_rate_mph] = results.at[idx, 'Effluent Gas Rate, mol/hr'] * results.at[idx, tag_norm] / 100.0

            out_gas_C_rate_mph += results.at[idx, tag_rate_mph] * component['C']
            out_gas_H_rate_mph += results.at[idx, tag_rate_mph] * component['H']
            out_gas_O_rate_mph += results.at[idx, tag_rate_mph] * component['O']
            out_gas_N_rate_mph += results.at[idx, tag_rate_mph] * component['N']
            
        results.at[idx, 'Effluent C Rate, mol/hr'] = out_gas_C_rate_mph
        results.at[idx, 'Effluent H Rate, mol/hr'] = out_gas_H_rate_mph
        results.at[idx, 'Effluent O Rate, mol/hr'] = out_gas_O_rate_mph
        results.at[idx, 'Effluent N Rate, mol/hr'] = out_gas_N_rate_mph

        results.at[idx, 'Effluent C Rate, g/hr'] = out_gas_C_rate_mph * 12.0
        results.at[idx, 'Effluent H Rate, g/hr'] = out_gas_H_rate_mph * 1.0
        results.at[idx, 'Effluent O Rate, g/hr'] = out_gas_O_rate_mph * 16.0
        results.at[idx, 'Effluent N Rate, g/hr'] = out_gas_N_rate_mph * 14.0

        ## Liquid Feed
        # Liquid feed rate is d(feed scale)/dt. This mass balance is designed for a single dilution rate (feed rate / V)
        # Step 1: check to see if rate at beginning and end of time period are the same
        t_start = start_time
        t_start_plus1 = start_time + 1.0
        t_end = end_time
        t_end_minus1 = end_time - 1.0

        # Feed rate = (feed scale reading @ t - feed scale reading @ t+1) / 1 hour
        # ACTUALLY in kg, NOT grams. ***Need to fix tag list and output files to say kilograms instead of grams
        feed_rate_start = raw_data.at[t_start, variable_dict['Feed Scale, kg']] - raw_data.at[t_start_plus1, variable_dict['Feed Scale, kg']]
        feed_rate_end = raw_data.at[t_end_minus1, variable_dict['Feed Scale, kg']] - raw_data.at[t_end, variable_dict['Feed Scale, kg']]

        # Conditional statement in case we don't have feed scale data
        if feed_rate_end == 0.0:
            # take user input for dilution rate
            assumed_liq_feed_rate = float(input("Feed scale data is missing. Please enter your assumed dilution rate, e.g. '0.03': ")) * assumed_working_mass_g
            results.at[idx, 'Liquid Feed Rate, g/hr'] = assumed_liq_feed_rate
        else:
            # scales all should be reading in kg units => convert to g
            results.at[idx, 'Liquid Feed Rate, g/hr'] = feed_rate_end * 1000.0

        # Dilution rate (will just be printed to output DF and excel workbook)
        results.at[idx, 'Dilution Rate, hrs^-1'] = results.at[idx, 'Liquid Feed Rate, g/hr'] / assumed_working_mass_g

        results.at[idx, 'Liquid C Rate, g/hr'] = results.at[idx, 'Liquid Feed Rate, g/hr'] * water_c_h_o_n_ratio_mass[0]
        results.at[idx, 'Liquid H Rate, g/hr'] = results.at[idx, 'Liquid Feed Rate, g/hr'] * water_c_h_o_n_ratio_mass[1]
        results.at[idx, 'Liquid O Rate, g/hr'] = results.at[idx, 'Liquid Feed Rate, g/hr'] * water_c_h_o_n_ratio_mass[2]
        results.at[idx, 'Liquid N Rate, g/hr'] = results.at[idx, 'Liquid Feed Rate, g/hr'] * water_c_h_o_n_ratio_mass[3]

        ## Product Generation
        # Start with accumulation. Accumulation of C = Mass * (delta_w_bdo * sigma_O_bdo + delta_w_bm * sigma_O_bm + delta_w_h2o * sigma_O_h2o)
        delta_w_bdo_gpL = (bdo_end - bdo_start) / 1000.0
        delta_w_bm_gpL = (od_end - od_end) * OD_to_gDCW_perL
        delta_w_h2o_gpL = -(delta_w_bdo_gpL + delta_w_bm_gpL)
        results.at[idx, 'Accumulation C, g/hr'] = (assumed_working_mass_g / 1000.0 ) * ( delta_w_bdo_gpL * bdo_c_h_o_n_ratio_mass[0] + \
                                                                                    delta_w_bm_gpL * biomass_c_h_o_n_ratio_mass[0] + \
                                                                                    delta_w_h2o_gpL * water_c_h_o_n_ratio_mass[0] ) \
                                            / (end_time - start_time)
        results.at[idx, 'Accumulation H, g/hr'] = (assumed_working_mass_g / 1000.0 ) * ( delta_w_bdo_gpL * bdo_c_h_o_n_ratio_mass[1] + \
                                                                                    delta_w_bm_gpL * biomass_c_h_o_n_ratio_mass[1] + \
                                                                                    delta_w_h2o_gpL * water_c_h_o_n_ratio_mass[1] ) \
                                            / (end_time - start_time)
        results.at[idx, 'Accumulation O, g/hr'] = (assumed_working_mass_g / 1000.0 ) * ( delta_w_bdo_gpL * bdo_c_h_o_n_ratio_mass[2] + \
                                                                                    delta_w_bm_gpL * biomass_c_h_o_n_ratio_mass[2] + \
                                                                                    delta_w_h2o_gpL * water_c_h_o_n_ratio_mass[2] ) \
                                            / (end_time - start_time)
        results.at[idx, 'Accumulation N, g/hr'] = (assumed_working_mass_g / 1000.0 ) * ( delta_w_bdo_gpL * bdo_c_h_o_n_ratio_mass[3] + \
                                                                                    delta_w_bm_gpL * biomass_c_h_o_n_ratio_mass[3] + \
                                                                                    delta_w_h2o_gpL * water_c_h_o_n_ratio_mass[3] ) \
                                            / (end_time - start_time)

        # Determine liquid mass fraction water
        results.at[idx, 'w_BDO_liq_g/L'] = results.at[idx, 'BDO, mg/L'] / 1000.0
        results.at[idx, 'w_BM_liq_g/L'] = results.at[idx, 'OD, raw'] * OD_to_gDCW_perL
        results.at[idx, 'w_H2O_liq_g/L'] = 1000.0 - results.at[idx, 'w_BDO_liq_g/L'] - results.at[idx, 'w_BM_liq_g/L']

        # Oxygen balance. In - Out = Acc, so m_in_O_gas + m_in_O_liq - m_acc_O = m_out_O_gas + m_out_O_liq
        # m_out_O_liq = v_out_liq * (w_BDO_liq * w_O_BDO + w_BM_liq * w_O_BDO + w_H2O_liq * w_O_H2O)
        # Another option is to simply set 'Draw Rate, g/hr' = 'Liquid Feed Rate, g/hr'
        results.at[idx, 'Draw Rate, L/hr'] = (results.at[idx, 'Liquid O Rate, g/hr'] + results.at[idx, 'Inlet O Rate, g/hr'] - results.at[idx, 'Effluent O Rate, g/hr'] - results.at[idx, 'Accumulation O, g/hr']) \
            / (results.at[idx, 'w_BDO_liq_g/L'] * bdo_c_h_o_n_ratio_mass[2] + results.at[idx, 'w_BM_liq_g/L'] * biomass_c_h_o_n_ratio_mass[2] + results.at[idx, 'w_H2O_liq_g/L'] * water_c_h_o_n_ratio_mass[2])
        results.at[idx, 'Draw Rate, g/hr'] = results.at[idx, 'Draw Rate, L/hr'] * 1000.0

        # Determine liquid draw elemental rates
        results.at[idx, 'Draw C Rate, g/hr'] = results.at[idx, 'Draw Rate, L/hr'] * (results.at[idx, 'w_BDO_liq_g/L'] * bdo_c_h_o_n_ratio_mass[0] + \
                                                                                            results.at[idx, 'w_BM_liq_g/L'] * biomass_c_h_o_n_ratio_mass[0] + \
                                                                                            results.at[idx, 'w_H2O_liq_g/L'] * water_c_h_o_n_ratio_mass[0])
        results.at[idx, 'Draw H Rate, g/hr'] = results.at[idx, 'Draw Rate, L/hr'] * (results.at[idx, 'w_BDO_liq_g/L'] * bdo_c_h_o_n_ratio_mass[1] + \
                                                                                            results.at[idx, 'w_BM_liq_g/L'] * biomass_c_h_o_n_ratio_mass[1] + \
                                                                                            results.at[idx, 'w_H2O_liq_g/L'] * water_c_h_o_n_ratio_mass[1])
        results.at[idx, 'Draw O Rate, g/hr'] = results.at[idx, 'Draw Rate, L/hr'] * (results.at[idx, 'w_BDO_liq_g/L'] * bdo_c_h_o_n_ratio_mass[2] + \
                                                                                            results.at[idx, 'w_BM_liq_g/L'] * biomass_c_h_o_n_ratio_mass[2] + \
                                                                                            results.at[idx, 'w_H2O_liq_g/L'] * water_c_h_o_n_ratio_mass[2])
        results.at[idx, 'Draw N Rate, g/hr'] = results.at[idx, 'Draw Rate, L/hr'] * (results.at[idx, 'w_BDO_liq_g/L'] * bdo_c_h_o_n_ratio_mass[3] + \
                                                                                            results.at[idx, 'w_BM_liq_g/L'] * biomass_c_h_o_n_ratio_mass[3] + \
                                                                                            results.at[idx, 'w_H2O_liq_g/L'] * water_c_h_o_n_ratio_mass[3])

        ## Calculate Closures
        results.at[idx, 'C Closure'] = ( results.at[idx, 'Effluent C Rate, g/hr'] + results.at[idx, 'Draw C Rate, g/hr'] + results.at[idx, 'Accumulation C, g/hr'] ) /\
                                        ( results.at[idx, 'Inlet C Rate, g/hr'] + results.at[idx, 'Liquid C Rate, g/hr'] )
        results.at[idx, 'H Closure'] = ( results.at[idx, 'Effluent H Rate, g/hr'] + results.at[idx, 'Draw H Rate, g/hr'] + results.at[idx, 'Accumulation H, g/hr'] ) /\
                                        ( results.at[idx, 'Inlet H Rate, g/hr'] + results.at[idx, 'Liquid H Rate, g/hr'] )
        results.at[idx, 'O Closure'] = ( results.at[idx, 'Effluent O Rate, g/hr'] + results.at[idx, 'Draw O Rate, g/hr'] + results.at[idx, 'Accumulation O, g/hr'] ) /\
                                        ( results.at[idx, 'Inlet O Rate, g/hr'] + results.at[idx, 'Liquid O Rate, g/hr'] )
        results.at[idx, 'N Closure'] = ( results.at[idx, 'Effluent N Rate, g/hr'] + results.at[idx, 'Draw N Rate, g/hr'] + results.at[idx, 'Accumulation N, g/hr'] ) /\
                                        ( results.at[idx, 'Inlet N Rate, g/hr'] + results.at[idx, 'Liquid N Rate, g/hr'] )

        print(f"Carbon Closure: {results.at[idx, 'C Closure']}")

        # Calculate EUR, CER, MUR, OUR, NUR

        KPIs = [{'rate':'CER', 'component':'carbon_dioxide'},
                #{'rate':'EUR', 'component':'ethane'}, 
                {'rate':'MUR', 'component':'methane'},
                {'rate':'NUR', 'component':'nitrogen'},
                {'rate':'OUR', 'component':'oxygen'}
        ]

        for KPI in KPIs:
            tag_out = 'F_' + KPI['component'] + '_out_molph'
            
            # Sum component molar flow rates through each MFC
            sum_F_in_molph = 0.0    
            for ref in used_MFCs:
                tag_in = f'{ref}_F_' + KPI['component'] + '_in_molph'
                sum_F_in_molph += results.at[idx, tag_in]
            
            # Output KPI * (1000 mmol/mol) to DF; CER = out - in, XUR = in - out 
            KPI_with_units = KPI['rate'] + ', mmol/hr'
            if KPI['rate'] == 'CER':
                results.at[idx, KPI_with_units] = (results.at[idx, tag_out] - sum_F_in_molph) * 1000
            else:
                results.at[idx, KPI_with_units] = (sum_F_in_molph - results.at[idx, tag_out]) * 1000

        # Calculate BiomassER and 23BDO ER
        Biomass_MW = 23.19
        _23BDO_MW = 90.14

        # Biomass ER: OD * (g DCW / L / OD) * L * (1/hr) / (g DCW / mol DCW) * (1000 mmol DCW / mol DCW) = mmol DCW / hr
        results.at[idx, 'Biomass ER, mmol/hr'] = results.at[idx, 'OD, raw'] * OD_to_gDCW_perL * \
                                                    results.at[idx, 'Assumed Working Vol, L'] / Biomass_MW * \
                                                    results.at[idx, 'Dilution Rate, hrs^-1'] * 1000

        # 23BDO ER: (mg BDO / L) * L * (1/hr) / (g BDO / mol DCW) * (1000 mmol BDO / mol BDO) * (g BDO / 1000 mg BDO) = mmol BDO / hr
        results.at[idx, '2,3-BDO, mmol/hr'] = results.at[idx, 'BDO, mg/L'] * \
                                                results.at[idx, 'Assumed Working Vol, L'] / _23BDO_MW * \
                                                results.at[idx, 'Dilution Rate, hrs^-1']

        # Re-organize headers so that Closure and KPI columns come first
        beginning_cols = ['Run_ID', 'Tank', 'Start Time, hours', 'End Time, hours', 
                        'C Closure', 'H Closure', 'O Closure', 'N Closure', 
                        'CER, mmol/hr', 'EUR, mmol/hr', 'MUR, mmol/hr', 'NUR, mmol/hr', 'OUR, mmol/hr', 
                        'Biomass ER, mmol/hr', '2,3-BDO, mmol/hr', 'DCW, g/L/OD', 'Assumed Working Vol, L'
        ]

        # Pull just this row, with re-organized cols, to append to outfile
        reorganized_df = results.loc[[idx], [col for col in beginning_cols] + [col for col in results if col not in beginning_cols]]

        # Exception handling for PermissionError (e.g. if someone has the output file open)
        try:
            # Append to first unoccupied row of 'Python Raw' worksheet in outfile
            with pd.ExcelWriter(outfile, engine='openpyxl', mode='a') as writer:
                book = load_workbook(outfile)
                writer.book = book
                writer.sheets = {ws.title: ws for ws in book.worksheets}
                reorganized_df.to_excel(writer, sheet_name='Python Raw', startrow=writer.sheets['Python Raw'].max_row, index=True, header=False)
        except PermissionError as permission_error:
            print("-----------------------------------------------------")
            print(permission_error)
            print("Please close the output file and try again!")

    except FileNotFoundError as file_error:
        print("-----------------------------------------------------")
        print(file_error)
        print("Make sure I/O filepaths are correct and try again!")

    print("-----------------------------------------------------")

    # Restart script upon user input
    while True:
        answer = input("Would you like to run again? Enter 'y' to run again, or 'n' to continue: ")
        if answer.lower() in ('y', 'n'):
            break
        print("Invalid input. Please enter 'y' or 'n'.")
    if answer.lower() == 'y':
        print("-----------------------------------------------------")
        continue
    else:
        break